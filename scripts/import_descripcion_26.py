#!/usr/bin/env python3
"""
Importar DESCRIPCION_26 desde Excel a base de datos SQLite.
Versión simplificada con openpyxl.
"""

import sqlite3
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import load_workbook
except ImportError:
    print("❌ ERROR: openpyxl no está instalado")
    print("   Instálalo con: pip install openpyxl")
    exit(1)

# Rutas
EXCEL_PATH = Path("/Volumes/WEBS/disano-scraper/data/output/Tarifa_API_26_FINAL_COMPLETO.xlsx")
DB_PATH = Path("/Volumes/WEBS/API_DISANO/database/tarifa_disano.db")

print("=" * 80)
print("IMPORTAR DESCRIPCION_26 DESDE EXCEL")
print("=" * 80)

# Leer Excel
print(f"\n📖 Leyendo Excel: {EXCEL_PATH}")
try:
    wb = load_workbook(EXCEL_PATH, read_only=True)
    ws = wb.active
    max_row = ws.max_row if ws.max_row else 0
    max_col = ws.max_column if ws.max_column else 0
    print(f"✅ Excel leído: {max_row} filas, {max_col} columnas")
except Exception as e:
    print(f"❌ ERROR leyendo Excel: {e}")
    exit(1)

# Encontrar columnas
print(f"\n📊 Buscando columnas...")
codigo_col = None
desc_col = None
desc_26_col = None

if ws[1]:  # Primera fila (headers)
    for idx, cell in enumerate(ws[1]):
        header = str(cell.value).strip().upper() if cell.value else ""
        print(f"   Columna {idx}: {header}")

        if 'CÓDIGO' in header or 'CODIGO' in header:
            codigo_col = idx + 1  # openpyxl es 1-based
        elif header == 'DESCRIPCION':
            desc_col = idx + 1
        elif header == 'DESCRIPCION_26':
            desc_26_col = idx + 1

print(f"\n✅ Columnas identificadas:")
print(f"   CÓDIGO: columna {codigo_col}")
print(f"   DESCRIPCION: columna {desc_col}")
print(f"   DESCRIPCION_26: columna {desc_26_col}")

if not codigo_col or not desc_col or not desc_26_col:
    print("\n❌ ERROR: No se encontraron todas las columnas necesarias")
    exit(1)

# Conectar a base de datos
print(f"\n💾 Conectando a base de datos: {DB_PATH}")
conn = sqlite3.connect(DB_PATH)
cursor = conn.cursor()

# Verificar que la columna existe
cursor.execute('PRAGMA table_info(productos);')
columns = [col[1] for col in cursor.fetchall()]

if 'DESCRIPCION_26' not in columns:
    print("\n❌ ERROR: La columna DESCRIPCION_26 no existe en la base de datos")
    print("   Ejecuta primero este comando SQL:")
    print('   ALTER TABLE productos ADD COLUMN "DESCRIPCION_26" TEXT;')
    conn.close()
    exit(1)
else:
    print("✅ Columna DESCRIPCION_26 existe en la base de datos")

# Importar DESCRIPCION_26
print(f"\n🚀 Importando DESCRIPCION_26...")
updated = 0
skipped = 0
errors = 0

# Comenzar desde la fila 2 (saltando header)
for row_idx in range(2, max_row + 1):
    try:
        codigo_cell = ws.cell(row=row_idx, column=codigo_col)
        descripcion_26_cell = ws.cell(row=row_idx, column=desc_26_col)

        codigo = codigo_cell.value if codigo_cell else None
        descripcion_26 = descripcion_26_cell.value if descripcion_26_cell else None

        if codigo:
            # Actualizar en base de datos
            cursor.execute('''
                UPDATE productos
                SET "DESCRIPCION_26" = ?
                WHERE [CÓDIGO] = ?
            ''', (descripcion_26, str(codigo)))
            updated += 1
        else:
            skipped += 1

    except Exception as e:
        errors += 1
        if errors <= 5:  # Mostrar solo primeros 5 errores
            print(f"  ❌ Error en fila {row_idx}: {e}")

# Confirmar cambios
conn.commit()

# Estadísticas
print(f"\n✅ Importación completada:")
print(f"   Filas actualizadas: {updated}")
print(f"   Filas saltadas (sin código): {skipped}")
print(f"   Errores: {errors}")

# Verificar datos importados
cursor.execute('''
    SELECT COUNT(*),
           SUM(CASE WHEN "DESCRIPCION_26" IS NOT NULL THEN 1 ELSE 0 END) as with_data,
           SUM(CASE WHEN "DESCRIPCION_26" IS NULL THEN 1 ELSE 0 END) as null_data
    FROM productos
''')
total, with_data, null_data = cursor.fetchone()

print(f"\n📊 Estadísticas en base de datos:")
print(f"   Total filas: {total}")
print(f"   Con DESCRIPCION_26: {with_data}")
print(f"   Sin DESCRIPCION_26 (NULL): {null_data}")

# Mostrar ejemplos
print(f"\n📋 Ejemplos de datos importados:")
cursor.execute('''
    SELECT [CÓDIGO], DESCRIPCION, "DESCRIPCION_26"
    FROM productos
    WHERE "DESCRIPCION_26" IS NOT NULL
    LIMIT 5
''')

print("   CÓDIGO      | DESCRIPCION                       | DESCRIPCION_26")
print("   " + "-" * 80)
for row in cursor.fetchall():
    codigo = row[0]
    desc = row[1]
    desc_26 = row[2]

    # Truncar si es muy largo
    desc_str = (desc[:40] + "...") if desc and len(str(desc)) > 40 else (str(desc) if desc else "NULL")
    desc_26_str = (desc_26[:40] + "...") if desc_26 and len(str(desc_26)) > 40 else (str(desc_26) if desc_26 else "NULL")

    print(f"   {str(codigo):<12} | {desc_str:<40} | {desc_26_str}")

conn.close()

print(f"\n✅ Proceso completado!")
print(f"\n🕐 Hora de finalización: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
