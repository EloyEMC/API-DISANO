#!/usr/bin/env python3
"""
Migration: Import url_imagen from Excel to database
Purpose: Add url_imagen column and import from Excel column 57
Date: 2026-03-21

Usage: python3 02_import_url_imagen.py
"""

import sqlite3
import openpyxl
from datetime import datetime
import os
import sys

# Configuration
DB_PATH = '/Volumes/WEBS/API_DISANO/database/tarifa_disano.db'
EXCEL_PATH = '/Volumes/WEBS/disano-scraper/data/output/Tarifa_API_26_FINAL_COMPLETO.xlsx'
EXCEL_COLUMN_INDEX = 57  # Column number (1-based)
EXCEL_COLUMN_HEADER = 'Url_imagen'
CODIGO_COLUMN_INDEX = 3  # FIXED: CÓDIGO is in column C (3), not column B (2)

def add_column_if_not_exists(conn):
    """Add url_imagen column if it doesn't exist"""
    cursor = conn.cursor()
    cursor.execute("PRAGMA table_info(productos)")
    columns = [row[1] for row in cursor.fetchall()]
    
    if 'url_imagen' not in columns:
        print(f"Adding column 'url_imagen' to productos table...")
        cursor.execute("ALTER TABLE productos ADD COLUMN url_imagen TEXT")
        conn.commit()
        print("✓ Column 'url_imagen' added successfully")
    else:
        print("✓ Column 'url_imagen' already exists (idempotent)")

def import_url_imagen_from_excel(conn):
    """Import url_imagen data from Excel column 57"""
    print(f"\nLoading Excel file: {EXCEL_PATH}")
    
    # Load Excel file
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
    except Exception as e:
        print(f"✗ Error loading Excel file: {e}")
        sys.exit(1)
    
    # Verify column header
    header = ws.cell(row=1, column=EXCEL_COLUMN_INDEX).value
    if header != EXCEL_COLUMN_HEADER:
        print(f"✗ Column mismatch: Expected '{EXCEL_COLUMN_HEADER}', got '{header}'")
        sys.exit(1)
    
    print(f"✓ Column verified: {EXCEL_COLUMN_HEADER} (column {EXCEL_COLUMN_INDEX})")
    print(f"✓ CÓDIGO column: column {CODIGO_COLUMN_INDEX}")
    print(f"✓ Excel rows: {ws.max_row - 1} (excluding header)")
    
    # Import data
    cursor = conn.cursor()
    updated_count = 0
    null_count = 0
    not_found_count = 0
    
    # Prepare batch update
    updates = []
    
    for row_idx in range(2, ws.max_row + 1):
        codigo = ws.cell(row=row_idx, column=CODIGO_COLUMN_INDEX).value  # FIXED: Column 3 is CÓDIGO
        url_imagen = ws.cell(row=row_idx, column=EXCEL_COLUMN_INDEX).value
        
        # Skip if código is missing
        if not codigo:
            continue
        
        # Prepare value (NULL if empty or None)
        value = url_imagen if url_imagen and str(url_imagen).strip() else None
        
        if value:
            updates.append((value, codigo))
        else:
            null_count += 1
    
    # Execute batch updates
    print(f"\nUpdating database with {len(updates)} records...")
    cursor.executemany(
        "UPDATE productos SET url_imagen = ? WHERE CÓDIGO = ?",
        updates
    )
    updated_count = cursor.rowcount
    
    conn.commit()
    
    print(f"✓ Updated {updated_count} records with url_imagen")
    print(f"✓ {null_count} records have empty/NULL url_imagen")
    
    # Check for records not found in database
    cursor.execute("SELECT COUNT(*) FROM productos WHERE url_imagen IS NULL")
    db_null_count = cursor.fetchone()[0]
    not_found_count = db_null_count - null_count
    
    if not_found_count > 0:
        print(f"⚠ {not_found_count} database records not found in Excel")

def verify_import(conn):
    """Verify the import was successful"""
    print("\n=== Verification ===")
    cursor = conn.cursor()
    
    # Count records
    cursor.execute("SELECT COUNT(*) FROM productos")
    total = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(*) FROM productos WHERE url_imagen IS NOT NULL")
    not_null = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(*) FROM productos WHERE url_imagen IS NULL")
    null_count = cursor.fetchone()[0]
    
    # Sample data
    cursor.execute("SELECT CÓDIGO, url_imagen FROM productos WHERE url_imagen IS NOT NULL LIMIT 3")
    samples = cursor.fetchall()
    
    print(f"Total records: {total}")
    print(f"Records with url_imagen: {not_null} ({not_null/total*100:.1f}%)")
    print(f"Records without url_imagen: {null_count} ({null_count/total*100:.1f}%)")
    
    print("\nSample records:")
    for codigo, url in samples:
        display_url = url[:80] + "..." if len(url) > 80 else url
        print(f"  [{codigo}] {display_url}")

def main():
    """Main execution"""
    print("=" * 70)
    print("Migration: Import url_imagen from Excel")
    print("=" * 70)
    
    # Check files exist
    if not os.path.exists(DB_PATH):
        print(f"✗ Database not found: {DB_PATH}")
        sys.exit(1)
    
    if not os.path.exists(EXCEL_PATH):
        print(f"✗ Excel file not found: {EXCEL_PATH}")
        sys.exit(1)
    
    # Connect to database
    print(f"\nConnecting to database: {DB_PATH}")
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = OFF")
    
    try:
        # Step 1: Add column if not exists
        add_column_if_not_exists(conn)
        
        # Step 2: Import from Excel
        import_url_imagen_from_excel(conn)
        
        # Step 3: Verify
        verify_import(conn)
        
        print("\n" + "=" * 70)
        print("✓ Migration completed successfully!")
        print("=" * 70)
        
    except Exception as e:
        print(f"\n✗ Error during migration: {e}")
        conn.rollback()
        sys.exit(1)
    finally:
        conn.close()

if __name__ == '__main__':
    main()
